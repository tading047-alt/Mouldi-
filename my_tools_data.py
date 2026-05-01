# -*- coding: utf-8 -*-
"""
================================================================================
مكتبة my_data_tools - أدوات متكاملة لمعالجة البيانات والإرسال إلى Telegram
================================================================================

هذه المكتبة تحتوي على جميع الدوال التالية:

📊 دوال التحويل (من ملفات إلى DataFrame):
   1. excel_to_dataframe      - تحويل ملف Excel إلى DataFrame
   2. csv_to_dataframe        - تحويل ملف CSV إلى DataFrame
   3. google_sheet_to_dataframe - تحويل Google Sheet إلى DataFrame

📝 دوال الكتابة (من DataFrame إلى ملفات):
   4. dataframe_to_excel      - كتابة DataFrame في ملف Excel
   5. dataframe_to_csv        - كتابة DataFrame في ملف CSV
   6. dataframe_to_pdf        - كتابة DataFrame في ملف PDF

📁 دوال إدارة مجلد output:
   7. update_output_excel     - تحديث مجلد output وكتابة Excel
   8. update_output_multi_excel - كتابة عدة DataFrames في ملف واحد
   9. save_with_timestamp     - حفظ مع إضافة تاريخ ووقت
  10. read_latest_excel       - قراءة آخر ملف تم حفظه
  11. clean_output_folder     - تنظيف مجلد output من الملفات القديمة

🎲 دوال التوليد العشوائي (للتجربة):
  12. generate_random_numbers     - توليد أرقام عشوائية
  13. generate_random_number      - توليد رقم عشوائي بتوزيعات مختلفة
  14. generate_random_dataframe   - توليد DataFrame عشوائي كامل
  15. generate_sales_data         - توليد بيانات مبيعات عشوائية
  16. generate_employees_data     - توليد بيانات موظفين عشوائية
  17. generate_students_data      - توليد بيانات طلاب عشوائية

🤖 دوال الإرسال إلى Telegram:
  18. send_telegram_message   - إرسال رسالة إلى Telegram
  19. send_csv_link           - إرسال رابط تحميل CSV
  20. send_excel_link         - إرسال رابط تحميل Excel
  21. send_pdf_link           - إرسال رابط تحميل PDF
  22. send_any_file_link      - إرسال رابط أي ملف

🛠️ دوال مساعدة:
  23. get_dataframe_info      - الحصول على معلومات عن DataFrame

================================================================================
المؤلف: Mouldi
الإصدار: 2.0.0
التاريخ: 2024
================================================================================
"""

# ============================================================================
# المكتبات المطلوبة
# ============================================================================

import pandas as pd
import os
import random
import numpy as np
import requests
import shutil
import glob
from datetime import datetime, timedelta
from io import StringIO, BytesIO

# محاولة استيراد fpdf لإنشاء PDF
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False
    print("⚠️ مكتبة fpdf2 غير مثبتة. استخدم: pip install fpdf2")

# محاولة استيراد openpyxl لقراءة Excel
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ مكتبة openpyxl غير مثبتة. استخدم: pip install openpyxl")

# محاولة استيراد مكتبات Google Sheets
try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False


# ============================================================================
# القسم 1: دوال التحويل (من ملفات إلى DataFrame)
# ============================================================================

def excel_to_dataframe(file_path, sheet_name=0, **kwargs):
    """
    تحويل ملف Excel إلى DataFrame
    
    المعاملات:
    -----------
    file_path : str
        مسار ملف Excel
    sheet_name : str or int, default=0
        اسم أو رقم الورقة
    
    المخرجات:
    ----------
    pandas.DataFrame
        البيانات المحولة
    """
    try:
        if not os.path.exists(file_path):
            print(f"❌ الملف غير موجود: {file_path}")
            return None
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
        print(f"✅ تم تحويل {file_path} إلى DataFrame بنجاح")
        print(f"   الشكل: {df.shape[0]} صف × {df.shape[1]} عمود")
        return df
    except Exception as e:
        print(f"❌ خطأ في تحويل الملف: {e}")
        return None


def csv_to_dataframe(file_path, encoding='utf-8-sig', **kwargs):
    """
    تحويل ملف CSV إلى DataFrame
    
    المعاملات:
    -----------
    file_path : str
        مسار ملف CSV
    encoding : str, default='utf-8-sig'
        ترميز الملف
    
    المخرجات:
    ----------
    pandas.DataFrame
        البيانات المحولة
    """
    try:
        if not os.path.exists(file_path):
            print(f"❌ الملف غير موجود: {file_path}")
            return None
        
        df = pd.read_csv(file_path, encoding=encoding, **kwargs)
        print(f"✅ تم تحويل {file_path} إلى DataFrame بنجاح")
        print(f"   الشكل: {df.shape[0]} صف × {df.shape[1]} عمود")
        return df
    except Exception as e:
        print(f"❌ خطأ في تحويل الملف: {e}")
        return None


def google_sheet_to_dataframe(sheet_url, sheet_name=None, credentials_file=None):
    """
    تحويل ملف Google Sheet إلى DataFrame
    
    المعاملات:
    -----------
    sheet_url : str
        رابط ملف Google Sheet
    sheet_name : str, optional
        اسم الورقة
    credentials_file : str, optional
        مسار ملف الاعتمادات JSON (للملفات الخاصة)
    """
    try:
        # استخراج معرف الملف من الرابط
        if '/d/' in sheet_url:
            file_id = sheet_url.split('/d/')[1].split('/')[0]
        elif 'spreadsheets/d/' in sheet_url:
            file_id = sheet_url.split('spreadsheets/d/')[1].split('/')[0]
        else:
            file_id = sheet_url
        
        # حالة الملفات العامة
        if credentials_file is None:
            export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv"
            response = requests.get(export_url)
            
            if response.status_code == 200:
                df = pd.read_csv(StringIO(response.text))
                print(f"✅ تم تحويل Google Sheet إلى DataFrame بنجاح")
                print(f"   الشكل: {df.shape[0]} صف × {df.shape[1]} عمود")
                return df
            else:
                print(f"❌ فشل تحميل الملف. تأكد من أن الملف عام")
                return None
        
        # حالة الملفات الخاصة
        else:
            if not GOOGLE_SHEETS_AVAILABLE:
                print("❌ مكتبات Google Sheets غير مثبتة")
                return None
            
            scope = ['https://spreadsheets.google.com/feeds', 
                    'https://www.googleapis.com/auth/drive']
            creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
            client = gspread.authorize(creds)
            sheet = client.open_by_key(file_id)
            worksheet = sheet.worksheet(sheet_name) if sheet_name else sheet.sheet1
            data = worksheet.get_all_values()
            
            if len(data) > 1:
                df = pd.DataFrame(data[1:], columns=data[0])
            else:
                df = pd.DataFrame()
            
            print(f"✅ تم تحويل Google Sheet (خاص) إلى DataFrame بنجاح")
            print(f"   الشكل: {df.shape[0]} صف × {df.shape[1]} عمود")
            return df
            
    except Exception as e:
        print(f"❌ خطأ في تحويل Google Sheet: {e}")
        return None


# ============================================================================
# القسم 2: دوال الكتابة (من DataFrame إلى ملفات)
# ============================================================================

def dataframe_to_excel(df, output_path, sheet_name='Sheet1', **kwargs):
    """
    كتابة DataFrame في ملف Excel
    """
    try:
        dir_path = os.path.dirname(output_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        
        df.to_excel(output_path, sheet_name=sheet_name, **kwargs)
        print(f"✅ تم حفظ DataFrame في ملف Excel: {output_path}")
        print(f"   عدد الصفوف: {len(df)}, عدد الأعمدة: {len(df.columns)}")
        return True
    except Exception as e:
        print(f"❌ خطأ في حفظ ملف Excel: {e}")
        return False


def dataframe_to_csv(df, output_path, encoding='utf-8-sig', **kwargs):
    """
    كتابة DataFrame في ملف CSV
    """
    try:
        dir_path = os.path.dirname(output_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        
        df.to_csv(output_path, encoding=encoding, **kwargs)
        print(f"✅ تم حفظ DataFrame في ملف CSV: {output_path}")
        print(f"   عدد الصفوف: {len(df)}, عدد الأعمدة: {len(df.columns)}")
        return True
    except Exception as e:
        print(f"❌ خطأ في حفظ ملف CSV: {e}")
        return False


def dataframe_to_pdf(df, output_path, title="تقرير البيانات", orientation='L', include_date=True):
    """
    كتابة DataFrame في ملف PDF
    """
    if not FPDF_AVAILABLE:
        print("❌ مكتبة fpdf2 غير مثبتة. استخدم: pip install fpdf2")
        return False
    
    try:
        if df.empty:
            print("❌ DataFrame فارغ، لا يمكن إنشاء PDF")
            return False
        
        dir_path = os.path.dirname(output_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        
        class PDFReport(FPDF):
            def header(self):
                if self.page_no() > 1:
                    self.set_font('Arial', 'B', 12)
                    self.cell(0, 10, title, 0, 1, 'C')
                    self.ln(5)
            
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'صفحة {self.page_no()}', 0, 0, 'C')
        
        pdf = PDFReport(orientation=orientation, format='A4')
        pdf.add_page()
        
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, title, 0, 1, 'C')
        pdf.ln(10)
        
        if include_date:
            pdf.set_font("Arial", "I", 10)
            pdf.cell(0, 10, f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, 'R')
            pdf.ln(5)
        
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, f"عدد الصفوف: {len(df)}", 0, 1)
        pdf.cell(0, 10, f"عدد الأعمدة: {len(df.columns)}", 0, 1)
        pdf.ln(5)
        
        page_width = pdf.w - 20
        col_width = min(page_width / len(df.columns), 60) if len(df.columns) > 0 else 190
        
        pdf.set_font("Arial", "B", 9)
        for col in df.columns:
            pdf.cell(col_width, 10, str(col)[:30], 1, 0, 'C')
        pdf.ln()
        
        pdf.set_font("Arial", "", 8)
        for idx, row in df.head(100).iterrows():
            for col in df.columns:
                cell_text = str(row[col])[:25]
                pdf.cell(col_width, 8, cell_text, 1, 0, 'L')
            pdf.ln()
        
        if len(df) > 100:
            pdf.set_font("Arial", "I", 9)
            pdf.cell(0, 10, f"... و {len(df) - 100} صف إضافي", 0, 1, 'C')
        
        pdf.output(output_path)
        print(f"✅ تم حفظ DataFrame في ملف PDF: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ خطأ في حفظ ملف PDF: {e}")
        return False


# ============================================================================
# القسم 3: دوال إدارة مجلد output
# ============================================================================

def update_output_excel(df, filename="data.xlsx", output_folder="output", 
                         sheet_name='Sheet1', create_backup=True, **kwargs):
    """
    تحديث مجلد output وكتابة DataFrame في ملف Excel
    """
    result = {'success': False, 'folder_path': '', 'file_path': '', 'backup_path': None, 'message': ''}
    
    try:
        if df is None or df.empty:
            result['message'] = "❌ DataFrame فارغ، لا يمكن الكتابة"
            print(result['message'])
            return result
        
        folder_path = os.path.join(os.getcwd(), output_folder)
        os.makedirs(folder_path, exist_ok=True)
        result['folder_path'] = folder_path
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        file_path = os.path.join(folder_path, filename)
        result['file_path'] = file_path
        
        if create_backup and os.path.exists(file_path):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"{filename.replace('.xlsx', '')}_backup_{timestamp}.xlsx"
            backup_path = os.path.join(folder_path, backup_filename)
            shutil.copy2(file_path, backup_path)
            result['backup_path'] = backup_path
            print(f"💾 تم إنشاء نسخة احتياطية: {backup_filename}")
        
        df.to_excel(file_path, sheet_name=sheet_name, **kwargs)
        
        result['success'] = True
        result['message'] = f"✅ تم حفظ DataFrame في {file_path}"
        
        print(f"\n{'='*50}")
        print(f"📁 تحديث مجلد output:")
        print(f"   📂 المجلد: {folder_path}")
        print(f"   📄 الملف: {filename}")
        print(f"   📊 الأبعاد: {df.shape[0]} صف × {df.shape[1]} عمود")
        print(f"{'='*50}")
        
        return result
        
    except Exception as e:
        result['message'] = f"❌ خطأ: {e}"
        print(result['message'])
        return result


def update_output_multi_excel(dataframes, filename="multi_data.xlsx", 
                               output_folder="output", create_backup=True):
    """
    كتابة عدة DataFrames في أوراق مختلفة من ملف Excel واحد
    """
    result = {'success': False, 'folder_path': '', 'file_path': '', 'sheets': [], 'message': ''}
    
    try:
        folder_path = os.path.join(os.getcwd(), output_folder)
        os.makedirs(folder_path, exist_ok=True)
        result['folder_path'] = folder_path
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        file_path = os.path.join(folder_path, filename)
        result['file_path'] = file_path
        
        if create_backup and os.path.exists(file_path):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(folder_path, f"backup_{timestamp}_{filename}")
            shutil.copy2(file_path, backup_path)
            print(f"💾 تم إنشاء نسخة احتياطية: backup_{timestamp}_{filename}")
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            if isinstance(dataframes, dict):
                for sheet_name, df in dataframes.items():
                    if df is not None and not df.empty:
                        df.to_excel(writer, sheet_name=str(sheet_name)[:31], index=False)
                        result['sheets'].append(sheet_name)
                        print(f"   ✅ كتابة ورقة: {sheet_name} ({df.shape[0]} صف)")
            elif isinstance(dataframes, list):
                for i, df in enumerate(dataframes):
                    if df is not None and not df.empty:
                        sheet_name = f"Sheet_{i+1}"
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        result['sheets'].append(sheet_name)
                        print(f"   ✅ كتابة ورقة: {sheet_name} ({df.shape[0]} صف)")
        
        result['success'] = True
        result['message'] = f"✅ تم حفظ {len(result['sheets'])} ورقة في {file_path}"
        
        print(f"\n📁 تم حفظ {len(result['sheets'])} ورقة في {filename}")
        return result
        
    except Exception as e:
        result['message'] = f"❌ خطأ: {e}"
        print(result['message'])
        return result


def save_with_timestamp(df, prefix="data", output_folder="output", sheet_name='Sheet1', **kwargs):
    """
    حفظ DataFrame مع إضافة تاريخ ووقت للاسم
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{prefix}_{timestamp}.xlsx"
    return update_output_excel(df, filename=filename, output_folder=output_folder,
                                sheet_name=sheet_name, create_backup=False, **kwargs)


def read_latest_excel(output_folder="output", pattern="*.xlsx"):
    """
    قراءة آخر ملف Excel تم حفظه في مجلد output
    """
    folder_path = os.path.join(os.getcwd(), output_folder)
    
    if not os.path.exists(folder_path):
        print(f"❌ مجلد {output_folder} غير موجود")
        return None
    
    excel_files = glob.glob(os.path.join(folder_path, pattern))
    
    if not excel_files:
        print(f"❌ لا توجد ملفات Excel في مجلد {output_folder}")
        return None
    
    latest_file = max(excel_files, key=os.path.getmtime)
    
    print(f"📂 آخر ملف: {os.path.basename(latest_file)}")
    print(f"   تاريخ التعديل: {datetime.fromtimestamp(os.path.getmtime(latest_file))}")
    
    try:
        excel_data = pd.ExcelFile(latest_file)
        sheets = excel_data.sheet_names
        
        if len(sheets) == 1:
            df = pd.read_excel(latest_file)
            print(f"   الأبعاد: {df.shape[0]} صف × {df.shape[1]} عمود")
            return df
        else:
            dataframes = {}
            for sheet in sheets:
                dataframes[sheet] = pd.read_excel(latest_file, sheet_name=sheet)
                print(f"   ورقة '{sheet}': {dataframes[sheet].shape[0]} صف")
            return dataframes
            
    except Exception as e:
        print(f"❌ خطأ في قراءة الملف: {e}")
        return None


def clean_output_folder(output_folder="output", days_old=30, pattern=None):
    """
    تنظيف مجلد output من الملفات القديمة
    """
    result = {'deleted_count': 0, 'deleted_files': [], 'folder_path': '', 'message': ''}
    
    folder_path = os.path.join(os.getcwd(), output_folder)
    result['folder_path'] = folder_path
    
    if not os.path.exists(folder_path):
        result['message'] = f"⚠️ مجلد {output_folder} غير موجود"
        print(result['message'])
        return result
    
    now = datetime.now()
    cutoff = now - timedelta(days=days_old)
    
    if pattern:
        search_pattern = os.path.join(folder_path, pattern)
        files = glob.glob(search_pattern)
    else:
        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path)]
        files = [f for f in files if os.path.isfile(f)]
    
    for file_path in files:
        file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
        if file_time < cutoff:
            os.remove(file_path)
            result['deleted_count'] += 1
            result['deleted_files'].append(os.path.basename(file_path))
    
    result['message'] = f"✅ تم حذف {result['deleted_count']} ملف قديم"
    print(result['message'])
    return result


# ============================================================================
# القسم 4: دوال التوليد العشوائي
# ============================================================================

def generate_random_numbers(count=10, min_value=0, max_value=100, decimal_places=0):
    """
    توليد قائمة بأرقام عشوائية
    """
    numbers = []
    for _ in range(count):
        if decimal_places == 0:
            num = random.randint(min_value, max_value)
        else:
            num = round(random.uniform(min_value, max_value), decimal_places)
        numbers.append(num)
    
    print(f"✅ تم توليد {count} رقم عشوائي")
    print(f"   المدى: [{min_value}, {max_value}]")
    return numbers


def generate_random_number(min_value=0, max_value=100, decimal_places=0, count=1, distribution='uniform'):
    """
    توليد رقم عشوائي بتوزيعات مختلفة
    """
    if distribution == 'uniform':
        if decimal_places == 0:
            numbers = [random.randint(min_value, max_value) for _ in range(count)]
        else:
            numbers = [round(random.uniform(min_value, max_value), decimal_places) for _ in range(count)]
    
    elif distribution == 'normal':
        mean = (min_value + max_value) / 2
        std = (max_value - min_value) / 6
        numbers = []
        for _ in range(count):
            num = np.random.normal(mean, std)
            num = max(min_value, min(num, max_value))
            numbers.append(round(num, decimal_places) if decimal_places > 0 else int(num))
    
    elif distribution == 'exponential':
        numbers = []
        scale = (max_value - min_value) / 3
        for _ in range(count):
            num = min_value + np.random.exponential(scale)
            num = min(num, max_value)
            numbers.append(round(num, decimal_places) if decimal_places > 0 else int(num))
    
    else:
        print(f"❌ توزيع غير معروف: {distribution}")
        return None
    
    if count == 1:
        return numbers[0]
    
    print(f"✅ تم توليد {count} رقم عشوائي (توزيع: {distribution})")
    return numbers


def generate_random_dataframe(rows=10, columns=None, random_seed=None):
    """
    توليد DataFrame عشوائي
    """
    if random_seed is not None:
        random.seed(random_seed)
        np.random.seed(random_seed)
    
    first_names = ['أحمد', 'محمد', 'علي', 'حسن', 'سارة', 'فاطمة', 'نورة', 'عائشة', 'خالد', 'عمر']
    last_names = ['المالكي', 'العتيبي', 'الدوسري', 'الشمري', 'الغامدي', 'الزهراني']
    arabic_cities = ['الرياض', 'جدة', 'مكة', 'المدينة', 'الدمام', 'الخبر', 'الطائف', 'تبوك']
    
    if columns is None:
        columns = {
            'id': 'int',
            'الاسم': 'name',
            'العمر': 'int',
            'المدينة': 'city',
            'الراتب': 'float',
            'تاريخ_التسجيل': 'date'
        }
    
    data = {}
    
    for col_name, col_type in columns.items():
        col_type = col_type.lower()
        
        if col_type == 'int':
            data[col_name] = [random.randint(18, 60) for _ in range(rows)]
        elif col_type == 'float':
            data[col_name] = [round(random.uniform(3000, 20000), 2) for _ in range(rows)]
        elif col_type == 'name':
            data[col_name] = [f"{random.choice(first_names)} {random.choice(last_names)}" for _ in range(rows)]
        elif col_type == 'city':
            data[col_name] = [random.choice(arabic_cities) for _ in range(rows)]
        elif col_type == 'date':
            end_date = datetime.now()
            start_date = end_date - timedelta(days=365 * 3)
            data[col_name] = [(start_date + timedelta(days=random.randint(0, 1095))).strftime('%Y-%m-%d') for _ in range(rows)]
        elif col_type == 'bool':
            data[col_name] = [random.choice([True, False]) for _ in range(rows)]
        else:
            data[col_name] = [f"قيمة_{random.randint(1, 1000)}" for _ in range(rows)]
    
    df = pd.DataFrame(data)
    print(f"✅ تم توليد DataFrame عشوائي: {rows} صف × {len(columns)} عمود")
    return df


def generate_sales_data(rows=20, start_date='2024-01-01', end_date='2024-12-31'):
    """
    توليد بيانات مبيعات عشوائية
    """
    products = ['لابتوب', 'هاتف', 'جهاز لوحي', 'سماعة', 'شاحن', 'ماوس', 'لوحة مفاتيح', 'طابعة', 'كاميرا']
    regions = ['الرياض', 'جدة', 'مكة', 'المدينة', 'الدمام', 'الخبر', 'تبوك', 'أبها']
    
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    date_range = (end - start).days
    
    data = {
        'رقم_الفاتورة': [f'INV-{i:05d}' for i in range(1, rows + 1)],
        'التاريخ': [(start + timedelta(days=random.randint(0, date_range))).strftime('%Y-%m-%d') for _ in range(rows)],
        'المنتج': [random.choice(products) for _ in range(rows)],
        'الكمية': [random.randint(1, 20) for _ in range(rows)],
        'السعر': [random.randint(50, 5000) for _ in range(rows)],
        'المنطقة': [random.choice(regions) for _ in range(rows)]
    }
    
    df = pd.DataFrame(data)
    df['الإجمالي'] = df['الكمية'] * df['السعر']
    
    print(f"✅ تم توليد {rows} سجل مبيعات عشوائي")
    print(f"   إجمالي الإيرادات: {df['الإجمالي'].sum():,.2f}")
    return df


def generate_employees_data(rows=15):
    """
    توليد بيانات موظفين عشوائية
    """
    departments = ['IT', 'HR', 'Finance', 'Marketing', 'Sales', 'Operations']
    positions = ['مدير', 'محلل', 'مبرمج', 'محاسب', 'مسوق', 'مندوب مبيعات']
    
    data = {
        'الموظف_ID': [f'EMP-{i:03d}' for i in range(1, rows + 1)],
        'الاسم': [random.choice(['أحمد', 'محمد', 'علي', 'سارة', 'فاطمة']) + ' ' + 
                  random.choice(['المالكي', 'العتيبي', 'الدوسري']) for _ in range(rows)],
        'العمر': [random.randint(22, 55) for _ in range(rows)],
        'القسم': [random.choice(departments) for _ in range(rows)],
        'المنصب': [random.choice(positions) for _ in range(rows)],
        'الراتب': [random.randint(4000, 25000) for _ in range(rows)]
    }
    
    df = pd.DataFrame(data)
    print(f"✅ تم توليد {rows} سجل موظفين عشوائي")
    return df


def generate_students_data(rows=20):
    """
    توليد بيانات طلاب عشوائية
    """
    majors = ['علوم حاسوب', 'هندسة', 'طب', 'إدارة أعمال', 'محاسبة']
    
    data = {
        'رقم_الطالب': [f'STU-{i:04d}' for i in range(1, rows + 1)],
        'الاسم': [random.choice(['أحمد', 'محمد', 'سارة', 'فاطمة', 'نورة']) + ' ' +
                  random.choice(['التميمي', 'الغامدي', 'الزهراني']) for _ in range(rows)],
        'المعدل': [round(random.uniform(2.0, 4.0), 2) for _ in range(rows)],
        'التخصص': [random.choice(majors) for _ in range(rows)]
    }
    
    df = pd.DataFrame(data)
    print(f"✅ تم توليد {rows} سجل طلاب عشوائي")
    return df


# ============================================================================
# القسم 5: دوال الإرسال إلى Telegram
# ============================================================================

def send_telegram_message(bot_token, chat_id, message, parse_mode='HTML'):
    """
    إرسال رسالة نصية إلى Telegram
    """
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {'chat_id': chat_id, 'text': message, 'parse_mode': parse_mode}
    
    try:
        response = requests.post(url, json=payload, timeout=30)
        if response.status_code == 200:
            print("✅ تم إرسال الرسالة إلى Telegram")
            return True
        else:
            print(f"❌ فشل إرسال الرسالة")
            return False
    except Exception as e:
        print(f"❌ خطأ: {e}")
        return False


def send_csv_link(bot_token, chat_id, file_url, caption=""):
    """إرسال رابط CSV إلى Telegram"""
    message = f"📁 **رابط تحميل ملف CSV**\n\n{file_url}"
    if caption:
        message += f"\n\n{caption}"
    return send_telegram_message(bot_token, chat_id, message)


def send_excel_link(bot_token, chat_id, file_url, caption=""):
    """إرسال رابط Excel إلى Telegram"""
    message = f"📊 **رابط تحميل ملف Excel**\n\n{file_url}"
    if caption:
        message += f"\n\n{caption}"
    return send_telegram_message(bot_token, chat_id, message)


def send_pdf_link(bot_token, chat_id, file_url, caption=""):
    """إرسال رابط PDF إلى Telegram"""
    message = f"📄 **رابط تحميل ملف PDF**\n\n{file_url}"
    if caption:
        message += f"\n\n{caption}"
    return send_telegram_message(bot_token, chat_id, message)


def send_any_file_link(bot_token, chat_id, file_url, file_name="ملف", caption=""):
    """إرسال رابط أي ملف إلى Telegram"""
    message = f"📎 **رابط تحميل {file_name}**\n\n{file_url}"
    if caption:
        message += f"\n\n{caption}"
    return send_telegram_message(bot_token, chat_id, message)


# ============================================================================
# القسم 6: دوال مساعدة
# ============================================================================

def get_dataframe_info(df):
    """
    الحصول على معلومات مفصلة عن DataFrame
    """
    if df is None or df.empty:
        return {'is_empty': True, 'rows': 0, 'columns': 0, 'column_names': []}
    
    return {
        'is_empty': False,
        'rows': len(df),
        'columns': len(df.columns),
        'column_names': df.columns.tolist(),
        'dtypes': df.dtypes.astype(str).to_dict(),
        'null_counts': df.isnull().sum().to_dict()
    }


# ============================================================================
# اختبار المكتبة (عند تشغيل الملف مباشرة)
# ============================================================================

if __name__ == "__main__":
    
    print("="*70)
    print("📦 مكتبة my_data_tools - اختبار جميع الدوال")
    print("="*70)
    
    # 1. اختبار دوال التوليد العشوائي
    print("\n" + "="*40)
    print("1. اختبار دوال التوليد العشوائي:")
    print("="*40)
    
    df_sales = generate_sales_data(10)
    df_employees = generate_employees_data(5)
    
    # 2. اختبار دوال مجلد output
    print("\n" + "="*40)
    print("2. اختبار دوال مجلد output:")
    print("="*40)
    
    update_output_excel(df_sales, filename='sales_test.xlsx', index=False)
    save_with_timestamp(df_employees, prefix='employees')
    
    # 3. اختبار دوال الكتابة المتعددة
    print("\n" + "="*40)
    print("3. اختبار الكتابة المتعددة:")
    print("="*40)
    
    sheets = {'المبيعات': df_sales, 'الموظفين': df_employees}
    update_output_multi_excel(sheets, filename='complete_report.xlsx')
    
    # 4. عرض النتائج
    print("\n" + "="*40)
    print("4. الملفات الناتجة:")
    print("="*40)
    
    output_path = os.path.join(os.getcwd(), 'output')
    if os.path.exists(output_path):
        for f in os.listdir(output_path):
            file_path = os.path.join(output_path, f)
            size = os.path.getsize(file_path) / 1024
            print(f"   📄 {f} - {size:.2f} KB")
    
    print("\n" + "="*70)
    print("🎉 جميع الدوال تعمل بنجاح!")
    print("="*70)
